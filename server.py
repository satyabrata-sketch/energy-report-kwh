import http.server
import socketserver
import json
import openpyxl
import io
import os
import ctypes
import datetime

PORT = 8080

def safe_copy_file(src, dst):
    """ Safely copy a file even if locked by Excel (EXCEL.EXE) on Windows """
    try:
        kernel32 = ctypes.windll.kernel32
        GENERIC_READ = 0x80000000
        FILE_SHARE_READ = 1
        FILE_SHARE_WRITE = 2
        FILE_SHARE_DELETE = 4
        OPEN_EXISTING = 3
        
        handle = kernel32.CreateFileW(
            src, GENERIC_READ,
            FILE_SHARE_READ | FILE_SHARE_WRITE | FILE_SHARE_DELETE,
            None, OPEN_EXISTING, 0, None
        )
        if handle == -1 or handle == 0 or handle == 0xFFFFFFFF:
            return False
        
        buf_size = 65536
        buf = ctypes.create_string_buffer(buf_size)
        bytes_read = ctypes.c_ulong(0)
        
        os.makedirs(os.path.dirname(dst), exist_ok=True)
        with open(dst, 'wb') as f_out:
            while True:
                res = kernel32.ReadFile(handle, buf, buf_size, ctypes.byref(bytes_read), None)
                if not res or bytes_read.value == 0:
                    break
                f_out.write(buf.raw[:bytes_read.value])
                
        kernel32.CloseHandle(handle)
        return True
    except Exception as e:
        print("Safe copy error:", e)
        return False

def ensure_master_templates():
    f1 = 'DT-3 Daily energy Consumption  Updated.xlsx'
    d1 = 'master_templates/DT3_master.xlsx'
    if os.path.exists(f1):
        safe_copy_file(f1, d1)

    f2 = 'AHU_Saved_Hours_and_Energy_Cost_Report.xlsx'
    d2 = 'master_templates/AHU_master.xlsx'
    if os.path.exists(f2):
        safe_copy_file(f2, d2)

def Math_max_0(val):
    try:
        v = float(val)
        return v if v > 0 else 0.0
    except:
        return 0.0

def scan_dt3_sheet_rows(ws):
    """ Dynamically locate exact reading and consumption row indices for any month sheet """
    meter_rows = {}
    cons_rows = {}
    
    # Readings section: rows 1 to 42
    for r in range(1, 42):
        c_val = str(ws.cell(r, 3).value or '').strip()
        if '540430038840' in c_val: meter_rows['eb_1'] = r
        elif '540430038845' in c_val: meter_rows['eb_2'] = r
        elif '540430038844' in c_val: meter_rows['eb_3'] = r
        elif '540430038821' in c_val: meter_rows['eb_4'] = r
        elif '540430038849' in c_val: meter_rows['eb_5'] = r
        elif '540430038646' in c_val or '540430038696' in c_val: meter_rows['eb_6'] = r
        elif '540430038843' in c_val: meter_rows['eb_7'] = r
        elif '540430038848' in c_val: meter_rows['eb_8'] = r
        elif c_val == 'AHU1': meter_rows['ahu_1'] = r
        elif c_val == 'AHU2': meter_rows['ahu_2'] = r
        elif c_val == 'AHU3': meter_rows['ahu_3'] = r
        elif c_val == 'AHU4': meter_rows['ahu_4'] = r
        elif 'AHU1 - BTU' in c_val: meter_rows['btu_1'] = r
        elif 'AHU2 - BTU' in c_val: meter_rows['btu_2'] = r
        elif 'AHU3 - BTU' in c_val: meter_rows['btu_3'] = r
        elif 'AHU4 - BTU' in c_val: meter_rows['btu_4'] = r
        
    # Consumptions section: rows 40 to max_row
    for r in range(40, ws.max_row + 1):
        c_val = str(ws.cell(r, 3).value or '').strip()
        b_val = str(ws.cell(r, 2).value or '').strip()
        lbl = c_val or b_val
        if '540430038840' in lbl: cons_rows['eb_1'] = r
        elif '540430038845' in lbl: cons_rows['eb_2'] = r
        elif '540430038844' in lbl: cons_rows['eb_3'] = r
        elif '540430038821' in lbl: cons_rows['eb_4'] = r
        elif '540430038849' in lbl: cons_rows['eb_5'] = r
        elif '540430038646' in lbl or '540430038696' in lbl: cons_rows['eb_6'] = r
        elif '540430038843' in lbl: cons_rows['eb_7'] = r
        elif '540430038848' in lbl: cons_rows['eb_8'] = r
        elif lbl == 'AHU1': cons_rows['ahu_1'] = r
        elif lbl == 'AHU2': cons_rows['ahu_2'] = r
        elif lbl == 'AHU3': cons_rows['ahu_3'] = r
        elif lbl == 'AHU4': cons_rows['ahu_4'] = r
        elif 'AHU1 - BTU' in lbl: cons_rows['btu_1'] = r
        elif 'AHU2 - BTU' in lbl: cons_rows['btu_2'] = r
        elif 'AHU3 - BTU' in lbl: cons_rows['btu_3'] = r
        elif 'AHU4 - BTU' in lbl: cons_rows['btu_4'] = r
        elif 'Total Power - EB Unit' in lbl: cons_rows['tot_eb'] = r
        elif 'Total Power- DG' in lbl or 'Total Power - DG' in lbl: cons_rows['tot_dg'] = r
        elif 'Total Power Cumulative' in lbl: cons_rows['tot_cum'] = r
        elif 'Total AHU Power Consumption' in lbl: cons_rows['tot_ahu'] = r
        elif 'Total AHU DG consumption' in lbl or 'Total AHU DG Consumption' in lbl: cons_rows['tot_ahu_dg'] = r
        elif 'Total BTU Consumption' in lbl: cons_rows['tot_btu'] = r

    return meter_rows, cons_rows

def ensure_month_sheet_in_dt3(wb, year, month):
    """ Ensures that a specific month sheet (e.g. Aug-2026) exists in DT3 workbook """
    month_names = ['', 'Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun', 'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec']
    m_str = month_names[month]
    target_names = [f"{m_str}-{year}", f"{m_str} {year}", f"{m_str}-{str(year)[2:]}", f"{m_str}{year}"]
    
    for sname in wb.sheetnames:
        for t in target_names:
            if sname.lower() == t.lower():
                return sname
                
    # Needs to create sheet by cloning July-2026 or latest sheet
    base_sheet_name = 'July-2026' if 'July-2026' in wb.sheetnames else wb.sheetnames[-1]
    source = wb[base_sheet_name]
    new_sheet_name = f"{m_str}-{year}"
    target = wb.copy_worksheet(source)
    target.title = new_sheet_name
    
    # Calculate days in month
    import calendar
    _, days_in_month = calendar.monthrange(year, month)
    
    for day in range(1, days_in_month + 1):
        col = 4 + day # Col E is 5 (day 1)
        d = datetime.date(year, month, day)
        day_name = d.strftime('%a')
        
        target.cell(4, col, datetime.datetime(year, month, day, 0, 0))
        target.cell(5, col, day_name)
        target.cell(40, col, datetime.datetime(year, month, day, 0, 0))
        target.cell(41, col, day_name)
        if target.max_row >= 84:
            target.cell(84, col, datetime.datetime(year, month, day, 0, 0))
            target.cell(85, col, day_name)
            
    if target.max_row >= 83:
        target.cell(83, 5, f"{m_str.upper()} - {year}    EV METER READING")
        
    return new_sheet_name

class EnergyTrackerHandler(http.server.SimpleHTTPRequestHandler):
    def do_POST(self):
        if self.path == '/api/export/kwh':
            self.handle_export_kwh()
        elif self.path == '/api/export/ahu':
            self.handle_export_ahu()
        elif self.path == '/api/save':
            self.handle_save_data()
        else:
            self.send_error(404, "Endpoint not found")

    def handle_save_data(self):
        content_length = int(self.headers.get('Content-Length', 0))
        post_data = self.rfile.read(content_length)
        try:
            payload = json.loads(post_data.decode('utf-8'))
            with open('user_data.json', 'w', encoding='utf-8') as f:
                json.dump(payload, f, indent=2)
            self.send_response(200)
            self.send_header('Content-Type', 'application/json')
            self.end_headers()
            self.wfile.write(json.dumps({"status": "success"}).encode('utf-8'))
        except Exception as e:
            self.send_error(500, f"Error saving data: {str(e)}")

    def handle_export_kwh(self):
        content_length = int(self.headers.get('Content-Length', 0))
        post_data = self.rfile.read(content_length)
        ensure_master_templates()

        template_file = 'master_templates/DT3_master.xlsx'
        if not os.path.exists(template_file):
            template_file = 'DT-3 Daily energy Consumption  Updated.xlsx'

        try:
            tracker_data = json.loads(post_data.decode('utf-8')) if content_length > 0 else {}
            wb = openpyxl.load_workbook(template_file)
            
            wb.calculation.calcMode = 'auto'
            wb.calculation.fullCalcOnLoad = True

            kwh_daily = tracker_data.get('kwh_daily', {})

            # Ensure sheets exist for all months present in kwh_daily
            for d_str in kwh_daily.keys():
                try:
                    p = d_str.split('-')
                    y, m = int(p[0]), int(p[1])
                    ensure_month_sheet_in_dt3(wb, y, m)
                except:
                    pass

            # Map all date columns across month sheets & set freeze panes
            date_to_sheet_col = {}
            sheet_row_maps = {}

            for sheetname in wb.sheetnames:
                if sheetname in ['Dashboard', 'summary']:
                    continue
                sheet = wb[sheetname]
                if hasattr(sheet, 'sheet_view') and sheet.sheet_view:
                    sheet.sheet_view.showGridLines = True

                sheet.freeze_panes = 'E6'

                m_rows, c_rows = scan_dt3_sheet_rows(sheet)
                sheet_row_maps[sheetname] = (m_rows, c_rows)

                for r in range(2, 6):
                    for c in range(5, sheet.max_column + 1):
                        val = sheet.cell(r, c).value
                        if isinstance(val, (datetime.datetime, datetime.date)):
                            d_str = val.strftime('%Y-%m-%d')
                            date_to_sheet_col[d_str] = (sheetname, c)

            # Update readings and consumptions for each date
            month_totals = {} # { "2026-08": { eb: 0, ahu: 0, dg: 0, btu: 0 } }

            for d_str, day_entry in kwh_daily.items():
                m_key_ym = d_str[:7] # e.g. "2026-08"
                if m_key_ym not in month_totals:
                    month_totals[m_key_ym] = {'eb': 0.0, 'ahu': 0.0, 'dg': 0.0, 'btu': 0.0}

                if d_str in date_to_sheet_col:
                    sname, c_idx = date_to_sheet_col[d_str]
                    sheet = wb[sname]
                    m_rows, c_rows = sheet_row_maps[sname]

                    # 1. EB Meters
                    eb_list = day_entry.get('eb', [])
                    day_eb_sum = 0.0
                    day_dg_sum = 0.0

                    for eb_item in eb_list:
                        m_id = eb_item.get('id')
                        m_key = f'eb_{m_id}'
                        if m_key in m_rows:
                            eb_row = m_rows[m_key]
                            dg_row = eb_row + 1
                            eb_cons_row = c_rows.get(m_key)
                            dg_cons_row = eb_cons_row + 1 if eb_cons_row else None

                            curr_eb = float(eb_item.get('reading', 0))
                            curr_dg = float(eb_item.get('dg_reading', 0))

                            prev_col = c_idx - 1
                            prev_eb = float(sheet.cell(eb_row, prev_col).value or 0)
                            prev_dg = float(sheet.cell(dg_row, prev_col).value or 0)

                            sheet.cell(eb_row, c_idx, curr_eb)
                            sheet.cell(dg_row, c_idx, curr_dg)

                            eb_cons = Math_max_0(curr_eb - prev_eb)
                            dg_cons = Math_max_0(curr_dg - prev_dg)

                            if eb_cons_row:
                                sheet.cell(eb_cons_row, c_idx, eb_cons)
                            if dg_cons_row:
                                sheet.cell(dg_cons_row, c_idx, dg_cons)

                            day_eb_sum += eb_cons
                            day_dg_sum += dg_cons

                    if 'tot_eb' in c_rows: sheet.cell(c_rows['tot_eb'], c_idx, day_eb_sum)
                    if 'tot_dg' in c_rows: sheet.cell(c_rows['tot_dg'], c_idx, day_dg_sum)
                    if 'tot_cum' in c_rows: sheet.cell(c_rows['tot_cum'], c_idx, day_eb_sum + day_dg_sum)

                    month_totals[m_key_ym]['eb'] += day_eb_sum
                    month_totals[m_key_ym]['dg'] += day_dg_sum

                    # 2. AHU Meters
                    ahu_list = day_entry.get('ahu', [])
                    day_ahu_sum = 0.0
                    day_ahu_dg_sum = 0.0

                    for ahu_item in ahu_list:
                        m_id = ahu_item.get('id')
                        m_key = f'ahu_{m_id}'
                        if m_key in m_rows:
                            ahu_row = m_rows[m_key]
                            dg_row = ahu_row + 1
                            ahu_cons_row = c_rows.get(m_key)
                            dg_cons_row = ahu_cons_row + 1 if ahu_cons_row else None

                            curr_ahu = float(ahu_item.get('reading', 0))
                            curr_dg = float(ahu_item.get('dg_reading', 0))

                            prev_col = c_idx - 1
                            prev_ahu = float(sheet.cell(ahu_row, prev_col).value or 0)
                            prev_dg = float(sheet.cell(dg_row, prev_col).value or 0)

                            sheet.cell(ahu_row, c_idx, curr_ahu)
                            sheet.cell(dg_row, c_idx, curr_dg)

                            ahu_cons = Math_max_0(curr_ahu - prev_ahu)
                            dg_cons = Math_max_0(curr_dg - prev_dg)

                            if ahu_cons_row:
                                sheet.cell(ahu_cons_row, c_idx, ahu_cons)
                            if dg_cons_row:
                                sheet.cell(dg_cons_row, c_idx, dg_cons)

                            day_ahu_sum += ahu_cons
                            day_ahu_dg_sum += dg_cons

                    if 'tot_ahu' in c_rows: sheet.cell(c_rows['tot_ahu'], c_idx, day_ahu_sum)
                    if 'tot_ahu_dg' in c_rows: sheet.cell(c_rows['tot_ahu_dg'], c_idx, day_ahu_dg_sum)

                    month_totals[m_key_ym]['ahu'] += day_ahu_sum

                    # 3. BTU Meters
                    btu_list = day_entry.get('btu', [])
                    day_btu_sum = 0.0

                    for btu_item in btu_list:
                        m_id = btu_item.get('id')
                        m_key = f'btu_{m_id}'
                        if m_key in m_rows:
                            btu_row = m_rows[m_key]
                            btu_cons_row = c_rows.get(m_key)

                            curr_btu = float(btu_item.get('reading', 0))
                            prev_col = c_idx - 1
                            prev_btu = float(sheet.cell(btu_row, prev_col).value or 0)

                            sheet.cell(btu_row, c_idx, curr_btu)

                            btu_cons = Math_max_0(curr_btu - prev_btu)
                            if btu_cons_row:
                                sheet.cell(btu_cons_row, c_idx, btu_cons)

                            day_btu_sum += btu_cons

                    if 'tot_btu' in c_rows: sheet.cell(c_rows['tot_btu'], c_idx, day_btu_sum)

                    month_totals[m_key_ym]['btu'] += day_btu_sum

            # Update 'summary' sheet with latest month totals (e.g. August 2026)
            if 'summary' in wb.sheetnames:
                ws_sum = wb['summary']
                # Check for Aug 2026 column in summary (Col 28)
                aug_k = '2026-08'
                if aug_k in month_totals and ws_sum.max_column >= 27:
                    tot_data = month_totals[aug_k]
                    aug_col = 28
                    ws_sum.cell(1, aug_col, datetime.datetime(2026, 8, 1, 0, 0))
                    ws_sum.cell(8, aug_col, datetime.datetime(2026, 8, 1, 0, 0))
                    
                    eb_val = tot_data['eb']
                    ahu_val = tot_data['ahu']
                    dg_val = tot_data['dg']
                    btu_val = tot_data['btu']
                    tot_kwh = eb_val + ahu_val + dg_val + btu_val

                    ws_sum.cell(2, aug_col, eb_val)
                    ws_sum.cell(3, aug_col, ahu_val)
                    ws_sum.cell(4, aug_col, dg_val)
                    ws_sum.cell(5, aug_col, btu_val)
                    ws_sum.cell(6, aug_col, tot_kwh)

                    # Cost breakdown (EB rate: 7.45, AHU rate: 7.45, DG rate: 33.85, BTU rate: 4.30)
                    eb_cost = eb_val * 7.45
                    ahu_cost = ahu_val * 7.45
                    dg_cost = dg_val * 33.85
                    btu_cost = btu_val * 4.30
                    tot_cost = eb_cost + ahu_cost + dg_cost + btu_cost

                    ws_sum.cell(9, aug_col, eb_cost)
                    ws_sum.cell(10, aug_col, ahu_cost)
                    ws_sum.cell(11, aug_col, dg_cost)
                    ws_sum.cell(12, aug_col, btu_cost)
                    ws_sum.cell(13, aug_col, tot_cost)

            output = io.BytesIO()
            wb.save(output)
            output.seek(0)
            file_data = output.read()

        except Exception as e:
            print("Fallback to raw master file due to:", e)
            with open(template_file, 'rb') as f_raw:
                file_data = f_raw.read()

        self.send_response(200)
        self.send_header('Content-Type', 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        self.send_header('Content-Disposition', 'attachment; filename="DT-3 Daily energy Consumption  Updated.xlsx"')
        self.send_header('Content-Length', str(len(file_data)))
        self.end_headers()
        self.wfile.write(file_data)

    def handle_export_ahu(self):
        content_length = int(self.headers.get('Content-Length', 0))
        post_data = self.rfile.read(content_length)
        ensure_master_templates()

        template_file = 'master_templates/AHU_master.xlsx'
        if not os.path.exists(template_file):
            template_file = 'AHU_Saved_Hours_and_Energy_Cost_Report.xlsx'

        try:
            tracker_data = json.loads(post_data.decode('utf-8')) if content_length > 0 else {}
            wb = openpyxl.load_workbook(template_file)
            
            wb.calculation.calcMode = 'auto'
            wb.calculation.fullCalcOnLoad = True

            ahu_saving = tracker_data.get('ahu_saving', {})

            # Update monthly sheets
            for sheetname in wb.sheetnames:
                sheet = wb[sheetname]
                if hasattr(sheet, 'sheet_view') and sheet.sheet_view:
                    sheet.sheet_view.showGridLines = True

                if sheetname in ['INDEX & CONTROL PANEL', 'Executive Dashboard']:
                    continue

                sheet.freeze_panes = 'C7'

                date_row_map = {}
                for r in range(7, sheet.max_row + 1):
                    val = sheet.cell(r, 1).value
                    if isinstance(val, (datetime.datetime, datetime.date)):
                        d_str = val.strftime('%Y-%m-%d')
                        date_row_map[d_str] = r

                for d_str, r_idx in date_row_map.items():
                    if d_str in ahu_saving:
                        day_entry = ahu_saving[d_str]
                        ahus = day_entry.get('ahus', [])

                        col_offsets = [3, 13, 23, 33] # C, M, W, AG
                        for i, offset in enumerate(col_offsets):
                            if i < len(ahus):
                                a_item = ahus[i]
                                if a_item.get('on_time'):
                                    sheet.cell(r_idx, offset, str(a_item['on_time']))
                                if a_item.get('off_time'):
                                    sheet.cell(r_idx, offset + 1, str(a_item['off_time']))
                                if a_item.get('kwh_cons') is not None:
                                    sheet.cell(r_idx, offset + 5, float(a_item['kwh_cons']))

            # Update Executive Dashboard if present
            if 'Executive Dashboard' in wb.sheetnames:
                ws_dash = wb['Executive Dashboard']
                # Calculate August 2026 summary from ahu_saving
                aug_entries = [v for k, v in ahu_saving.items() if k.startswith('2026-08')]
                if aug_entries:
                    aug_sched = 0
                    aug_run = 0.0
                    aug_saved_ahus = [0.0, 0.0, 0.0, 0.0]
                    aug_kwh = 0.0
                    aug_cost_saved = 0.0

                    for entry in aug_entries:
                        ahus = entry.get('ahus', [])
                        for i, a in enumerate(ahus):
                            on_t = str(a.get('on_time', ''))
                            off_t = str(a.get('off_time', ''))
                            kwh_c = float(a.get('kwh_cons') or 0)
                            aug_kwh += kwh_c

                            if on_t and off_t and on_t != 'OFF' and off_t != 'OFF':
                                try:
                                    on_h, on_m = map(int, on_t.split(':'))
                                    off_h, off_m = map(int, off_t.split(':'))
                                    run_h = max(0, (off_h + off_m/60) - (on_h + on_m/60))
                                    saved_h = max(0, 20 - (off_h + off_m/60))
                                    aug_run += run_h
                                    if i < 4:
                                        aug_saved_ahus[i] += saved_h
                                    rate_hr = (kwh_c * 7.45) / run_h if run_h > 0 else 0
                                    aug_cost_saved += saved_h * rate_hr
                                except:
                                    pass

                    # Row 14 is August 2026
                    ws_dash.cell(14, 4, aug_run)
                    ws_dash.cell(14, 5, aug_saved_ahus[0])
                    ws_dash.cell(14, 6, aug_saved_ahus[1])
                    ws_dash.cell(14, 7, aug_saved_ahus[2])
                    ws_dash.cell(14, 8, aug_saved_ahus[3])
                    ws_dash.cell(14, 9, sum(aug_saved_ahus))
                    ws_dash.cell(14, 10, aug_kwh)
                    ws_dash.cell(14, 11, aug_kwh * 7.45)
                    ws_dash.cell(14, 12, aug_kwh * 7.45)
                    ws_dash.cell(14, 13, (aug_kwh * 7.45) + aug_cost_saved)
                    ws_dash.cell(14, 14, aug_cost_saved)

            output = io.BytesIO()
            wb.save(output)
            output.seek(0)
            file_data = output.read()

        except Exception as e:
            print("Fallback to raw master file due to:", e)
            with open(template_file, 'rb') as f_raw:
                file_data = f_raw.read()

        self.send_response(200)
        self.send_header('Content-Type', 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        self.send_header('Content-Disposition', 'attachment; filename="AHU_Saved_Hours_and_Energy_Cost_Report.xlsx"')
        self.send_header('Content-Length', str(len(file_data)))
        self.end_headers()
        self.wfile.write(file_data)

if __name__ == '__main__':
    ensure_master_templates()
    handler = EnergyTrackerHandler
    with socketserver.TCPServer(("", PORT), handler) as httpd:
        print(f"Energy Tracker Server running on port {PORT} with real-time export engine...")
        httpd.serve_forever()
